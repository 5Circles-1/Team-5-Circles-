#!/usr/bin/env python3
"""Builds "5 Circles HQ" — the team's one operating sheet.

Output: dist/5-circles-hq.xlsx, built to be uploaded to Google Drive where it
converts to a native Google Sheet. That Sheet IS the tool: formulas, dropdowns
and conditional formatting only — no scripts, no macros, nothing to authorize,
nothing to host, nothing that can break.

Google Sheets is the runtime, so the file is written Google-native:
  - The 📌 Today lists are four spill formulas (FILTER / SORT / TAKE / HSTACK,
    stored with their Excel `_xlfn` names so Drive's xlsx converter accepts them)
    reading the data tabs. Counters are plain COUNTIFS. The counter logic and
    every seed expectation were first proven in a classic INDEX/MATCH build
    verified under LibreOffice (see git history), then the list mechanics were
    swapped for native spills; the deployed sheet is verified by reading its
    computed values back from Google after upload.
  - Formatting rides on column-level styles plus formatted seed rows (Google
    extends the row above when people type below), keeping the file tiny —
    small enough to travel as base64 through the Drive connector.
  - Today is sheet-protected: it fills itself; people type in the data tabs.

Palette comes from the 5 Circles design system (paper / ink / navy / Varsity
blue / teal / clay). Seed dates are relative to the build date so the dashboard
demonstrates itself on first open.
"""

import datetime as dt
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- palette --
NAVY = "0E1B2A"   # dark surface — headers
INK = "14181D"    # body text
SLATE = "6D7681"  # meta text
PAPER3 = "E7E3D9"
BLUE = "387ED1"   # the one functional accent
BLUE8 = "1F5BA0"  # blue for headings on paper
TEAL = "2F9C8E"   # good / done
CLAY = "C2543A"   # late / stuck
WHITE = "FFFFFF"

TEAL_TINT = "E3F1ED"
CLAY_TINT = "F7E4DE"
BLUE_TINT = "E4EDF8"

FONT = "Arial"

# ------------------------------------------------------------- sheet names --
S_TODAY = "📌 Today"
S_TASKS = "✅ Tasks"
S_CONTENT = "🎬 Content"
S_LEADS = "📞 Leads"
S_CLOSE = "🌙 Day Close"
S_TEAM = "👥 Team"
S_GUIDE = "📖 Read me"

def q(name):
    return "'" + name + "'"

N_ROWS = 1000
TEAM_ROWS = 51         # Team!A2:A51 feeds the name dropdowns

TODAY = dt.date.today()
D = lambda days: TODAY + dt.timedelta(days=days)

# ---------------------------------------------------------------- helpers --
def base_font(**kw):
    return Font(name=FONT, size=kw.pop("size", 10), color=kw.pop("color", INK), **kw)

def fill(color):
    return PatternFill("solid", fgColor=color)

def style_header_row(ws, headers, widths, tab_color):
    """Row 1: navy band with white bold headers. Freeze it. Set widths."""
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = base_font(bold=True, color=WHITE)
        c.fill = fill(NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

def style_col(ws, col, number_format=None, wrap=False, color=None):
    """Column-level default style (openpyxl ColumnDimension is styleable)."""
    cd = ws.column_dimensions[col]
    if number_format:
        cd.number_format = number_format
    cd.alignment = Alignment(vertical="top", wrap_text=wrap)
    cd.font = base_font(color=color or INK)

def seed_rows(ws, rows, date_cols=(), text_cols=(), slate_cols=()):
    """Write example rows with explicit per-cell formats so Google's
    format-follows-the-row-above behaviour propagates them."""
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            if v is None or v == "":
                continue
            c = ws.cell(row=i, column=j, value=v)
            col = get_column_letter(j)
            c.font = base_font(color=SLATE if col in slate_cols else INK)
            c.alignment = Alignment(vertical="top", wrap_text=(j == 1 or col in slate_cols))
            if col in date_cols:
                c.number_format = "dd-mmm"
            if col in text_cols:
                c.number_format = "@"

def dv_list(ws, formula, ranges):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    for r in ranges:
        dv.add(r)

def add_defined_name(wb, name, ref):
    dn = DefinedName(name, attr_text=ref)
    try:
        wb.defined_names.add(dn)          # openpyxl <= 3.0
    except AttributeError:
        wb.defined_names[name] = dn       # openpyxl >= 3.1

# ================================================================== build ==
wb = Workbook()
wb.remove(wb.active)

ws_today = wb.create_sheet(S_TODAY)
ws_tasks = wb.create_sheet(S_TASKS)
ws_content = wb.create_sheet(S_CONTENT)
ws_leads = wb.create_sheet(S_LEADS)
ws_close = wb.create_sheet(S_CLOSE)
ws_team = wb.create_sheet(S_TEAM)
ws_guide = wb.create_sheet(S_GUIDE)

add_defined_name(wb, "TeamNames", f"{q(S_TEAM)}!$A$2:$A${TEAM_ROWS}")

TQ, CQ, LQ = q(S_TASKS), q(S_CONTENT), q(S_LEADS)

# ------------------------------------------------------------------ Team ---
style_header_row(ws_team,
                 ["Name", "Role", "Phone", "Filed day close yesterday?"],
                 [20, 22, 16, 24], SLATE)
style_col(ws_team, "A")
style_col(ws_team, "B")
style_col(ws_team, "C", number_format="@")
style_col(ws_team, "D", color=SLATE)
seed_rows(ws_team, [("Rahul Sarawagi", "Founder", "", None)], text_cols=("C",))
ws_team["A2"].font = base_font(bold=True)
for r in range(2, TEAM_ROWS + 1):
    ws_team[f"D{r}"] = (
        f'=IF($A{r}="","",IF(COUNTIFS({q(S_CLOSE)}!$A:$A,TODAY()-1,'
        f'{q(S_CLOSE)}!$B:$B,$A{r})>0,"✔ yes","— no"))'
    )
    ws_team[f"D{r}"].font = base_font(color=SLATE)
ws_team.conditional_formatting.add(
    f"D2:D{TEAM_ROWS}",
    FormulaRule(formula=['LEFT($D2,1)="✔"'], font=base_font(color=TEAL)))

# ----------------------------------------------------------------- Tasks ---
style_header_row(ws_tasks,
                 ["Task", "Who", "Due", "Status", "Given by", "Given on",
                  "Done on", "Notes"],
                 [46, 14, 11, 11, 14, 11, 11, 34], BLUE8)
style_col(ws_tasks, "A", wrap=True)
style_col(ws_tasks, "B"); style_col(ws_tasks, "D"); style_col(ws_tasks, "E")
style_col(ws_tasks, "C", number_format="dd-mmm")
style_col(ws_tasks, "F", number_format="dd-mmm")
style_col(ws_tasks, "G", number_format="dd-mmm")
style_col(ws_tasks, "H", wrap=True, color=SLATE)

dv_list(ws_tasks, '"To do,Doing,Done,Stuck"', [f"D2:D{N_ROWS}"])
dv_list(ws_tasks, "TeamNames", [f"B2:B{N_ROWS}", f"E2:E{N_ROWS}"])

body = f"A2:H{N_ROWS}"
ws_tasks.conditional_formatting.add(
    body, FormulaRule(formula=['$D2="Done"'], fill=fill(TEAL_TINT),
                      font=base_font(color=SLATE), stopIfTrue=True))
ws_tasks.conditional_formatting.add(
    body, FormulaRule(formula=['AND($C2<>"",$C2<TODAY(),$D2<>"Done",$A2<>"")'],
                      fill=fill(CLAY_TINT)))
ws_tasks.conditional_formatting.add(
    body, FormulaRule(formula=['AND($C2<>"",$C2=TODAY(),$D2<>"Done",$A2<>"")'],
                      fill=fill(BLUE_TINT)))
ws_tasks.conditional_formatting.add(
    f"D2:D{N_ROWS}",
    FormulaRule(formula=['$D2="Stuck"'], font=base_font(color=CLAY, bold=True)))

seed_rows(ws_tasks, [
    ("Edit & schedule Thursday's theta reel", "Rahul Sarawagi", D(-1), "Doing",
     "Rahul Sarawagi", D(-2), None, "(sample rows — replace with real work)"),
    ("WhatsApp today's new ad leads before 1 pm", "Rahul Sarawagi", D(0), "To do",
     "Rahul Sarawagi", D(0), None, ""),
    ("Fix classroom mic echo before Saturday session", "Rahul Sarawagi", D(-2), "Stuck",
     "Rahul Sarawagi", D(-4), None, "Vendor not answering — need a decision"),
    ("Post Monday's market-myth carousel", "Rahul Sarawagi", D(-3), "Done",
     "Rahul Sarawagi", D(-5), D(-3), ""),
], date_cols=("C", "F", "G"), slate_cols=("H",))

# --------------------------------------------------------------- Content ---
style_header_row(ws_content,
                 ["Post date", "Account", "Format", "Topic / hook", "Stage",
                  "Owner", "Link", "Notes"],
                 [11, 14, 11, 42, 10, 14, 18, 26], TEAL)
style_col(ws_content, "A", number_format="dd-mmm")
style_col(ws_content, "B"); style_col(ws_content, "C")
style_col(ws_content, "D", wrap=True)
style_col(ws_content, "E"); style_col(ws_content, "F")
style_col(ws_content, "G", color=BLUE8)
style_col(ws_content, "H", wrap=True, color=SLATE)

dv_list(ws_content, '"5 Circles,Rahul,Traders Club"', [f"B2:B{N_ROWS}"])
dv_list(ws_content, '"Reel,Carousel,Story,Post,Live"', [f"C2:C{N_ROWS}"])
dv_list(ws_content, '"Idea,Script,Shoot,Edit,Ready,Posted"', [f"E2:E{N_ROWS}"])
dv_list(ws_content, "TeamNames", [f"F2:F{N_ROWS}"])

body = f"A2:H{N_ROWS}"
ws_content.conditional_formatting.add(
    body, FormulaRule(formula=['$E2="Posted"'], fill=fill(TEAL_TINT),
                      font=base_font(color=SLATE), stopIfTrue=True))
ws_content.conditional_formatting.add(
    body, FormulaRule(formula=['AND($A2<>"",$A2<TODAY(),$E2<>"Posted",$D2<>"")'],
                      fill=fill(CLAY_TINT)))
ws_content.conditional_formatting.add(
    body, FormulaRule(formula=['AND($A2<>"",$A2=TODAY(),$E2<>"Posted",$D2<>"")'],
                      fill=fill(BLUE_TINT)))

seed_rows(ws_content, [
    (D(1), "5 Circles", "Reel", "Theta: the rent you pay to hold hope", "Edit",
     "Rahul Sarawagi", "", "(sample rows — replace with your calendar)"),
    (D(-1), "Traders Club", "Post", "Saturday session recap + photos", "Shoot",
     "Rahul Sarawagi", "", ""),
    (D(3), "Rahul", "Carousel", "What my first loss taught me about sizing", "Idea",
     "Rahul Sarawagi", "", ""),
], date_cols=("A",), slate_cols=("H",))

# ----------------------------------------------------------------- Leads ---
style_header_row(ws_leads,
                 ["Added on", "Name", "Phone", "Source", "Program", "Status",
                  "Next follow-up", "Notes"],
                 [11, 18, 16, 12, 14, 12, 13, 30], CLAY)
style_col(ws_leads, "A", number_format="dd-mmm")
style_col(ws_leads, "B")
style_col(ws_leads, "C", number_format="@")
style_col(ws_leads, "D"); style_col(ws_leads, "E"); style_col(ws_leads, "F")
style_col(ws_leads, "G", number_format="dd-mmm")
style_col(ws_leads, "H", wrap=True, color=SLATE)

dv_list(ws_leads, '"Instagram,Ad,Webinar,Referral,Walk-in,Other"', [f"D2:D{N_ROWS}"])
dv_list(ws_leads, '"Options Lab,Traders Club,Workshop,Other"', [f"E2:E{N_ROWS}"])
dv_list(ws_leads, '"New,Contacted,Interested,Joined,Not now"', [f"F2:F{N_ROWS}"])

body = f"A2:H{N_ROWS}"
ws_leads.conditional_formatting.add(
    body, FormulaRule(formula=['$F2="Joined"'], fill=fill(TEAL_TINT), stopIfTrue=True))
ws_leads.conditional_formatting.add(
    body, FormulaRule(formula=['$F2="Not now"'], font=base_font(color=SLATE),
                      stopIfTrue=True))
ws_leads.conditional_formatting.add(
    body, FormulaRule(
        formula=['AND($B2<>"",$G2<>"",$G2<=TODAY(),$F2<>"Joined",$F2<>"Not now")'],
        fill=fill(CLAY_TINT)))

seed_rows(ws_leads, [
    (D(-1), "Aman Verma", "98765 43210", "Ad", "Options Lab", "New", D(0),
     "(sample rows — replace with real leads)"),
    (D(-3), "Priya S.", "91234 56780", "Webinar", "Traders Club", "Interested", D(2),
     "Wants Saturday batch details"),
], date_cols=("A", "G"), text_cols=("C",), slate_cols=("H",))
ws_leads["B2"].font = base_font(bold=True)
ws_leads["B3"].font = base_font(bold=True)

# ------------------------------------------------------------- Day Close ---
style_header_row(ws_close,
                 ["Date", "Name", "What I finished today",
                  "Stuck on (blank if nothing)", "Tomorrow's #1 thing"],
                 [11, 16, 44, 30, 30], NAVY)
style_col(ws_close, "A", number_format="dd-mmm")
style_col(ws_close, "B")
style_col(ws_close, "C", wrap=True)
style_col(ws_close, "D", wrap=True)
style_col(ws_close, "E", wrap=True)
dv_list(ws_close, "TeamNames", [f"B2:B{N_ROWS}"])
ws_close.conditional_formatting.add(
    f"A2:E{N_ROWS}",
    FormulaRule(formula=['AND($A2<>"",$A2=TODAY())'], fill=fill(BLUE_TINT)))
ws_close.conditional_formatting.add(
    f"D2:D{N_ROWS}",
    FormulaRule(formula=['AND($D2<>"",$A2<>"")'],
                font=base_font(color=CLAY, bold=True)))

seed_rows(ws_close, [
    (D(-1), "Rahul Sarawagi",
     "Recorded 2 reels, replied to 14 DMs, closed 1 admission",
     "", "Call Aman about Options Lab"),
], date_cols=("A",))

# ----------------------------------------------------------------- Today ---
wt = ws_today
wt.sheet_properties.tabColor = BLUE
for col, w in zip("ABCDEF", [30, 15, 12, 13, 6, 6]):
    wt.column_dimensions[col].width = w
wt.sheet_view.showGridLines = False
style_col(wt, "C", number_format="dd-mmm")

def band(row, text, size, height):
    wt.merge_cells(f"A{row}:D{row}")
    c = wt[f"A{row}"]
    c.value = text
    c.font = base_font(bold=(row == 1), color=WHITE if row == 1 else PAPER3, size=size)
    c.alignment = Alignment(vertical="center")
    for cc in "ABCD":
        wt[f"{cc}{row}"].fill = fill(NAVY)
    wt.row_dimensions[row].height = height

band(1, "5 CIRCLES HQ", 18, 32)
band(2, '="Live view · "&TEXT(TODAY(),"dddd, d mmm")&" · this page fills itself"', 10, 18)

def kpi(label_row, col, label, formula, color=INK):
    lc = wt[f"{col}{label_row}"]
    lc.value = label
    lc.font = base_font(size=8, color=SLATE, bold=True)
    nc = wt[f"{col}{label_row + 1}"]
    nc.value = formula
    nc.font = base_font(size=20, color=color, bold=True)
    nc.alignment = Alignment(vertical="center", horizontal="left")

kpi(4, "A", "OVERDUE TASKS",
    f'=COUNTIFS({TQ}!$C:$C,"<"&TODAY(),{TQ}!$D:$D,"<>Done",{TQ}!$A:$A,"<>")')
kpi(4, "B", "DUE TODAY",
    f'=COUNTIFS({TQ}!$C:$C,TODAY(),{TQ}!$D:$D,"<>Done",{TQ}!$A:$A,"<>")')
kpi(4, "C", "STUCK", f'=COUNTIF({TQ}!$D:$D,"Stuck")')
kpi(6, "A", "LEADS TO CALL",
    f'=COUNTIFS({LQ}!$G:$G,"<="&TODAY(),{LQ}!$F:$F,"<>Joined",'
    f'{LQ}!$F:$F,"<>Not now",{LQ}!$B:$B,"<>")', color=BLUE8)
kpi(6, "B", "CONTENT LATE",
    f'=COUNTIFS({CQ}!$A:$A,"<"&TODAY(),{CQ}!$E:$E,"<>Posted",{CQ}!$D:$D,"<>")')
kpi(6, "C", "CLOSED YDAY",
    f'=COUNTIFS({q(S_CLOSE)}!$A:$A,TODAY()-1)&" / "&COUNTA({q(S_TEAM)}!$A$2:$A${TEAM_ROWS})',
    color=SLATE)

for cell in ("A5", "B5", "C5", "B7"):
    wt.conditional_formatting.add(
        cell, CellIsRule(operator="greaterThan", formula=["0"],
                         font=base_font(size=20, bold=True, color=CLAY)))
    wt.conditional_formatting.add(
        cell, CellIsRule(operator="equal", formula=["0"],
                         font=base_font(size=20, bold=True, color=TEAL)))

def section(row, text, color, heads):
    wt.merge_cells(f"A{row}:D{row}")
    c = wt[f"A{row}"]
    c.value = text
    c.font = base_font(bold=True, color=color, size=12)
    c.alignment = Alignment(vertical="center")
    for cc in "ABCD":
        wt[f"{cc}{row}"].border = Border(bottom=Side(style="medium", color=color))
    wt.row_dimensions[row].height = 24
    for cc, lab in zip("ABCD", heads):
        h = wt[f"{cc}{row + 1}"]
        h.value = lab
        h.font = base_font(size=8, color=SLATE, bold=True)

# Each list below is ONE spill formula, capped with ARRAY_CONSTRAIN so it can
# never grow into the section beneath it. The KPI cells double as the
# empty-state test.
def col_range(sheet_q, col):
    return f"{sheet_q}!${col}$2:${col}${N_ROWS}"

tA, tB, tC, tD, tE = (col_range(TQ, c) for c in "ABCDE")
lB, lC, lE, lF, lG = (col_range(LQ, c) for c in "BCEFG")
cA, cB, cD, cE = (col_range(CQ, c) for c in "ABDE")

section(9, "⚠️  FIX FIRST — overdue & stuck", CLAY, ["TASK", "WHO", "DUE", "STATUS"])
wt["A11"] = (
    f'=IFERROR(_xlfn.TAKE(_xlfn._xlws.SORT(_xlfn._xlws.FILTER(_xlfn.HSTACK({tA},{tB},{tC},{tD}),'
    f'({tA}<>"")*(({tD}="Stuck")+({tC}<>"")*({tC}<TODAY())*({tD}<>"Done")>0)),3,TRUE),8),'
    f'IF($A$5+$C$5=0,"Nothing overdue, nothing stuck  ✅",""))'
)

section(20, "🔵  DUE TODAY", BLUE8, ["TASK", "WHO", "DUE", "GIVEN BY"])
wt["A22"] = (
    f'=IFERROR(_xlfn.TAKE(_xlfn._xlws.FILTER(_xlfn.HSTACK({tA},{tB},{tC},{tE}),'
    f'({tA}<>"")*({tC}=TODAY())*({tD}<>"Done")),6),'
    f'IF($B$5=0,"Nothing due today  ✅",""))'
)

section(29, "📞  CALL TODAY — follow-ups waiting", BLUE8,
        ["NAME", "PHONE", "PROGRAM", "STATUS"])
wt["A31"] = (
    f'=IFERROR(_xlfn.TAKE(_xlfn._xlws.FILTER(_xlfn.HSTACK({lB},{lC},{lE},{lF}),'
    f'({lB}<>"")*({lG}<>"")*({lG}<=TODAY())*({lF}<>"Joined")*({lF}<>"Not now")),6),'
    f'IF($A$7=0,"No follow-ups pending  ✅",""))'
)

section(38, "🎬  CONTENT RUNNING LATE", CLAY,
        ["TOPIC", "ACCOUNT", "PLANNED", "STAGE"])
wt["A40"] = (
    f'=IFERROR(_xlfn.TAKE(_xlfn._xlws.SORT(_xlfn._xlws.FILTER(_xlfn.HSTACK({cD},{cB},{cA},{cE}),'
    f'({cD}<>"")*({cA}<>"")*({cA}<TODAY())*({cE}<>"Posted")),3,TRUE),5),'
    f'IF($B$7=0,"Content is on schedule  ✅",""))'
)

for a in ("A11", "A22", "A31", "A40"):
    wt[a].font = base_font()
    wt[a].alignment = Alignment(vertical="top", wrap_text=True)

wt.conditional_formatting.add(
    "D11:D18", FormulaRule(formula=['$D11="Stuck"'],
                           font=base_font(color=CLAY, bold=True)))

wt.merge_cells("A46:D46")
c = wt["A46"]
c.value = ("Add work in ✅ Tasks · log every lead in 📞 Leads · file 🌙 Day Close "
           "before you leave. This page fills itself — don't type here.")
c.font = base_font(size=9, color=SLATE)
c.alignment = Alignment(vertical="center", wrap_text=True)
wt.row_dimensions[46].height = 28

wt.protection.sheet = True          # Today is read-only; data lives elsewhere

# ----------------------------------------------------------------- Guide ---
wg = ws_guide
wg.sheet_properties.tabColor = NAVY
wg.sheet_view.showGridLines = False
wg.column_dimensions["A"].width = 3
wg.column_dimensions["B"].width = 96

guide = [
    ("5 CIRCLES HQ — how this works", 16, NAVY, True, 30),
    ("One sheet. No app, no login, no server, nothing to install. It lives in your "
     "Google Drive, opens in the Sheets app on every phone, and never expires.", 10, INK, False, 30),
    ("", 10, INK, False, None),
    ("THE DAILY LOOP", 11, BLUE8, True, None),
    ("Morning — open 📌 Today. It shows what's overdue, what's due, which leads to "
     "call, and what content is late. Fix that list first.", 10, INK, False, 28),
    ("During the day — when work moves, tap its Status dropdown: To do → Doing → Done. "
     "Truly blocked? Set it to Stuck — that's not failure, that's a flare. Stuck rows "
     "jump to the top of Today so someone helps you fast.", 10, INK, False, 40),
    ("Evening — before you leave, add one row in 🌙 Day Close: what you finished, "
     "what you're stuck on, tomorrow's #1. Under a minute. Then send a ✅ in the "
     "WhatsApp group so everyone knows you closed.", 10, INK, False, 40),
    ("", 10, INK, False, None),
    ("THE TABS", 11, BLUE8, True, None),
    ("📌 Today — fills itself, and it's locked so nobody breaks it. Don't type here.", 10, INK, False, None),
    ("✅ Tasks — one row per job. A task without a Due date and a Who is just a wish.", 10, INK, False, None),
    ("🎬 Content — every post across 5 Circles, Rahul, and Traders Club. Move the "
     "Stage as it progresses: Idea → Script → Shoot → Edit → Ready → Posted.", 10, INK, False, 28),
    ("📞 Leads — every enquiry, same day it arrives. Set 'Next follow-up' and it will "
     "appear on Today when the day comes. A lead never chased is money burned on ads.", 10, INK, False, 28),
    ("🌙 Day Close — one row per person per day. The founder reads this tab first.", 10, INK, False, None),
    ("👥 Team — add your people here FIRST (name, role, phone). Every 'Who' dropdown "
     "in the sheet reads from this list.", 10, INK, False, 28),
    ("", 10, INK, False, None),
    ("RULES THAT KEEP IT ALIVE", 11, BLUE8, True, None),
    ("1.  If it's not in the sheet, it doesn't exist. WhatsApp is for talking; "
     "this is for remembering.", 10, INK, False, None),
    ("2.  Never delete rows — mark tasks Done, leads Not now. History is an asset.", 10, INK, False, None),
    ("3.  Colours mean things: red tint = late, blue tint = today, green tint = done.", 10, INK, False, None),
    ("4.  The sample rows (marked in Notes) are safe to replace with real work.", 10, INK, False, None),
    ("5.  Anyone can see everything. That's the point — problems can't hide.", 10, INK, False, None),
    ("", 10, INK, False, None),
    ("SETUP — 2 MINUTES, ONCE", 11, BLUE8, True, None),
    ("1.  👥 Team → add everyone's name, role, phone.", 10, INK, False, None),
    ("2.  Share (top-right) → add each person's Google email as Editor.", 10, INK, False, None),
    ("3.  Everyone installs the Google Sheets app and stars ⭐ this file.", 10, INK, False, None),
    ("That's the whole deployment.", 10, SLATE, False, None),
]
for r, (text, size, color, bold, height) in enumerate(guide, start=2):
    c = wg[f"B{r}"]
    c.value = text
    c.font = base_font(size=size, color=color, bold=bold)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    if height:
        wg.row_dimensions[r].height = height

# ------------------------------------------------------------------ save ---
os.makedirs("dist", exist_ok=True)
wb.calculation.fullCalcOnLoad = True
out = "dist/5-circles-hq.xlsx"
wb.save(out)
print(f"wrote {out} ({os.path.getsize(out)} bytes)")
